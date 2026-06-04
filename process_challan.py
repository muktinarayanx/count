import pdfplumber
import openpyxl
import sys
import io
import re

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# Full extraction from all pages
all_data = []

with pdfplumber.open("modified_all 10.pdf") as pdf:
    for i, page in enumerate(pdf.pages):
        text = page.extract_text()
        if not text:
            continue
        
        # Extract challan number
        challan_match = re.search(r'चालान\s*/?\s*पास\s*नं\.\s*:\s*(\d{18,25})', text)
        challan_no = challan_match.group(1) if challan_match else None
        
        # Extract weight (field 15 - mineral weight)
        weight_match = re.search(r'15.*?(\d+\.?\d*)\s*Mt', text)
        weight = None
        if weight_match:
            val = float(weight_match.group(1))
            if val < 1000:  # Filter out e-cap values
                weight = val
        
        if not weight:
            # Fallback: find small Mt values
            all_mt = re.findall(r'(\d+\.?\d*)\s*Mt', text)
            for v in all_mt:
                fv = float(v)
                if fv < 1000:
                    weight = fv
                    break
        
        if challan_no and weight is not None:
            all_data.append({
                'challan': challan_no,
                'weight': weight,
                'page': i + 1
            })
            print(f"Page {i+1}: Challan={challan_no}, Weight={weight} Mt")
        else:
            print(f"Page {i+1}: FAILED - Challan={challan_no}, Weight={weight}")

print(f"\nTotal records extracted: {len(all_data)}")

# Now write to Excel
print("\n--- Writing to Excel ---")
wb = openpyxl.load_workbook("challan check formula 1.xlsx")
ws = wb['Sheet1']

for idx, d in enumerate(all_data):
    row = idx + 2  # Data starts at row 2
    ws[f'A{row}'] = d['challan']
    ws[f'B{row}'] = d['weight']
    print(f"Row {row}: A={d['challan']}, B={d['weight']}")

# Save output
output_file = "challan_data_output.xlsx"
wb.save(output_file)
print(f"\nSaved to {output_file}")
print("Done!")
