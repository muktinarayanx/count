import http.server
import socketserver
import os
import re
import json
import pdfplumber
import openpyxl
from io import BytesIO

PORT = int(os.environ.get('PORT', 8080))


def parse_multipart(content_type, body):
    """Parse multipart/form-data without the cgi module."""
    # Extract boundary from content-type header
    boundary = None
    for part in content_type.split(';'):
        part = part.strip()
        if part.startswith('boundary='):
            boundary = part[len('boundary='):]
            break

    if not boundary:
        return [], None

    boundary_bytes = ('--' + boundary).encode()
    end_boundary = (boundary_bytes + b'--')

    parts = body.split(boundary_bytes)
    pdf_files = []
    excel_template = None

    for part in parts:
        part = part.strip(b'\r\n')
        if not part or part == b'--' or part == b'':
            continue

        # Split headers and body
        header_end = part.find(b'\r\n\r\n')
        if header_end == -1:
            continue

        header_section = part[:header_end].decode('utf-8', errors='replace')
        file_body = part[header_end + 4:]
        # Remove trailing \r\n
        if file_body.endswith(b'\r\n'):
            file_body = file_body[:-2]

        # Parse Content-Disposition
        filename = None
        field_name = None
        for line in header_section.split('\r\n'):
            if 'Content-Disposition' in line:
                for token in line.split(';'):
                    token = token.strip()
                    if token.startswith('filename="'):
                        filename = token[len('filename="'):-1]
                    elif token.startswith('name="'):
                        field_name = token[len('name="'):-1]

        if filename:
            if filename.lower().endswith('.pdf'):
                pdf_files.append((filename, file_body))
            elif filename.lower().endswith('.xlsx'):
                excel_template = file_body

    return pdf_files, excel_template


class ChallanHandler(http.server.SimpleHTTPRequestHandler):
    def end_headers(self):
        # Add CORS headers for all responses
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        super().end_headers()

    def do_OPTIONS(self):
        self.send_response(200)
        self.end_headers()

    def do_POST(self):
        if self.path == '/process':
            content_length = int(self.headers.get('Content-Length', 0))
            content_type = self.headers.get('Content-Type', '')
            body = self.rfile.read(content_length)

            # Parse multipart form data
            pdf_files, excel_template = parse_multipart(content_type, body)

            if not pdf_files:
                self.send_error(400, "No PDF files uploaded")
                return

            # Extract data from PDFs
            all_data = []
            for name, pdf_bytes in pdf_files:
                try:
                    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
                        for i, page in enumerate(pdf.pages):
                            text = page.extract_text()
                            if not text:
                                continue

                            # 1. Challan extraction
                            challan_match = re.search(r'\b\d{18,25}\b', text)
                            challan_no = challan_match.group(0) if challan_match else None

                            # 2. Weight extraction
                            weight = None
                            mt_matches = re.findall(r'(\d+\.?\d*)\s*[Mm]\s*[Tt]', text)
                            small_weights = [float(v) for v in mt_matches if 0 < float(v) < 1000]
                            if small_weights:
                                weight = small_weights[0]

                            if challan_no and weight is not None:
                                all_data.append({
                                    'challan': challan_no,
                                    'weight': weight,
                                    'page': i + 1,
                                    'fileName': name
                                })
                except Exception as e:
                    print(f"Error processing {name}: {e}")

            if not all_data:
                self.send_error(404, "No challan data found in the uploaded PDFs")
                return

            # Open template or create new workbook
            if excel_template:
                wb = openpyxl.load_workbook(BytesIO(excel_template))
            else:
                template_path = os.path.join(os.path.dirname(__file__), "challan check formula 1.xlsx")
                if os.path.exists(template_path):
                    wb = openpyxl.load_workbook(template_path)
                else:
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "Sheet1"
                    ws['A1'] = "unique no"
                    ws['B1'] = "Mt"

            ws = wb[wb.sheetnames[0]]
            start_row = 2

            # Write data
            for idx, d in enumerate(all_data):
                row = start_row + idx
                ws.cell(row=row, column=1, value=d['challan'])
                ws.cell(row=row, column=2, value=d['weight'])

            # Save workbook to memory
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            # Send response
            self.send_response(200)
            self.send_header('Content-Type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', 'attachment; filename="challan_data_output.xlsx"')
            self.end_headers()
            self.wfile.write(output.read())
        else:
            super().do_POST()


# Start server
print(f"Starting server on port {PORT}...")
socketserver.TCPServer.allow_reuse_address = True
with socketserver.TCPServer(("0.0.0.0", PORT), ChallanHandler) as httpd:
    print(f"Serving at http://0.0.0.0:{PORT}")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        pass
